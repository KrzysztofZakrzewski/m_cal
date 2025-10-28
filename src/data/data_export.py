import io
import streamlit as st
from openpyxl import Workbook
import pandas as pd
from fpdf import FPDF
from datetime import datetime
from io import BytesIO

# ===============================================================
# 💾 CSV Dowland functions
# ===============================================================
def download_csv_button(df, file_name, label="💾 Pobierz dane jako CSV"):
    """
    Helper function to generate a CSV download button in Streamlit.
    - df: DataFrame to export
    - file_name: output file name (with or without .csv)
    - label: button label
    """
    if not file_name.endswith(".csv"):
        file_name += ".csv"

    buffer = io.StringIO()
    df.to_csv(buffer, index=False)
    csv_data = buffer.getvalue()

    st.download_button(
        label=label,
        data=csv_data,
        file_name=file_name,
        mime="text/csv"
    )

# ===============================================================
# 💾 Exel Dowland functions
# ===============================================================
def to_excel(df: pd.DataFrame) -> bytes:
    """
    Convert a pandas DataFrame into an Excel file stored in memory as bytes.

    Parameters
    ----------
    df : pd.DataFrame
        The DataFrame to export. Must not be None or empty.

    Returns
    -------
    bytes
        The Excel file content in bytes, ready for download or saving to disk.

    Raises
    ------
    ValueError
        If the DataFrame is None or empty.
    
    Notes
    -----
    - Uses 'openpyxl' engine to write Excel.
    - Index column is excluded.
    - Sheet is named 'Dane'.
    """
    if df is None or df.empty:
        raise ValueError("❌ DataFrame is empty — nothing to export to Excel.")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Dane")
    processed_data = output.getvalue()
    return processed_data


# ===============================================================
# 💾 PDF Export functions for traing plan
# ===============================================================

def save_training_plan_to_pdf(text: str, subtitle: str = "Plan Treningowy") -> str:
    """
    Creates a PDF with the given text and saves it to a file.

    Args:
        text (str): Text content to include in the PDF.
        filename (str): Name of the PDF file to save.

    Returns:
        str: Path to the saved PDF file.
    """
    pdf = FPDF()
    pdf.add_page()
    # add a font that supports UTF-8
    pdf.add_font("DejaVu", "", "src/fonts/DejaVuSans.ttf", uni=True)
    pdf.add_font("DejaVu", "B", "src/fonts/DejaVuSans-Bold.ttf", uni=True)
    pdf.add_font("DejaVu", "I", "src/fonts/DejaVuSans-Oblique.ttf", uni=True)
    
    # Timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d")
    
    # Title
    pdf.set_font("DejaVu", "B", 16)
    pdf.cell(0, 10, "Plan Treningowy Dla:", ln=True, align="C")

    # Subtitle (second argument)
    pdf.set_font("DejaVu", "I", 12)
    pdf.cell(0, 10, f"{subtitle}, dnia: {timestamp}", ln=True, align="C")
    pdf.ln(10)

    # Main text
    pdf.set_font("DejaVu", "", 12)
    pdf.multi_cell(0, 8, text)

    # Add timestamp
    pdf.ln(10)
    pdf.set_font("DejaVu", "I", 10)
    pdf.cell(0, 10, f"📅 Wygenerowano: {timestamp}", align="R")

    # Save file (dynamic name)
    pdf_bytes = BytesIO()
    pdf.output(pdf_bytes)
    pdf_bytes.seek(0)  # wracamy na początek

    pdf__training_plan_filename = f"{subtitle.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d')}.pdf"

    return pdf_bytes, pdf__training_plan_filename